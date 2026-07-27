"""Print a compact, auditable formula map for the supplied LIFT workbooks."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


SOURCE_DIR = Path(r"C:\Users\Vadim\Downloads\Telegram Desktop")
WORKBOOK_NAMES = (
    "Lift 2ух стекло (2).xlsx",
    "Lift 2ух стеклопакет (2).xlsx",
    "Lift 3ех стекло  (2).xlsx",
    "Lift 3ех стеклопакет (2).xlsx",
    "Lift 4ех- стекло.xlsx",
    "Lift_4ех_стекло_глух_вверху_и_внизу (2).xlsx",
    "Lift 4ех- стеклопакет.xlsx",
    "Lift_4ех_стеклопакет_глух_вверху_и_внизу.xlsx",
)


def display(formula_cell, value_cell) -> str:
    formula = formula_cell.value
    if isinstance(formula, str) and formula.startswith("="):
        return f"{formula} => {value_cell.value!r}"
    return str(formula)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("needle", nargs="?", default="")
    args = parser.parse_args()
    selected = [
        filename
        for filename in WORKBOOK_NAMES
        if not args.needle or args.needle.casefold() in filename.casefold()
    ]

    for filename in selected:
        formula_book = load_workbook(SOURCE_DIR / filename, data_only=False)
        value_book = load_workbook(SOURCE_DIR / filename, data_only=True)
        formula_sheet = formula_book.worksheets[0]
        value_sheet = value_book[formula_sheet.title]

        print(f"\n## {filename}")
        for row in range(1, min(formula_sheet.max_row, 36) + 1):
            cells: list[str] = []
            for column in range(1, min(formula_sheet.max_column, 7) + 1):
                formula_cell = formula_sheet.cell(row, column)
                if formula_cell.value is None:
                    continue
                cells.append(
                    f"{formula_cell.coordinate}="
                    f"{display(formula_cell, value_sheet[formula_cell.coordinate])}"
                )
            if cells:
                print(" | ".join(cells))

        formula_book.close()
        value_book.close()


if __name__ == "__main__":
    main()
